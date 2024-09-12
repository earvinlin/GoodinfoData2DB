"""
20240901-1018 格式化產生寫入stocks_per_and_pbr的SQL Command
"""
import sys
import os 
import time
import openpyxl
import time
import platform
import re
from sqlalchemy import null
from datetime import datetime

NULL_VALUE = "null"
insertCnt = 0 
isFirstLine = True
try:
	print("[FormatFile.py] 開始執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

	if len(sys.argv) < 2 :
		print("You need input one parameter(資料日期, fmt : yyyymmdd)")
		print("syntax(windows)    : C:\python FormatStocksBzPerformanceData_perAndPbr.py 20240829PAP")
		print("syntax(imac/linux) : $python3 FormatStocksBzPerformanceData_perAndPbr.py 20240829PAP")
		sys.exit()

	theDate = sys.argv[1]

	# 設定檔案存取路徑
	loadFileDir = ""
	saveFileDir = ""
	if platform.system() == "Windows" :
		loadFileDir = "Data\\EXCEL\\Transfer\\bzPerformance\\" + str(theDate) + "\\"
		saveFileDir = "Data\\TXT\\bzPerformance\\" + str(theDate) + "\\"
	else :
		loadFileDir = "./Data/EXCEL/Transfer/bzPerformance/" + str(theDate) + "/"
		saveFileDir = "./Data/TXT/bzPerformance/" + str(theDate) + "/"

	print("loadFile dir: " + loadFileDir)
	print("saveFile dir: " + saveFileDir)

	# 20220522 取得要處理的檔案資料
	files = os.listdir(loadFileDir)
	# 以迴圈處理
	for inputFile in files:
		relativePath = os.path.join(loadFileDir, inputFile)
		if os.path.isfile(relativePath) :
			print("檔案：", inputFile)

		stockCode = inputFile[:4]
		outputFile = stockCode + ".txt"
		print("outfile: " + outputFile)
		outfile = open(saveFileDir + outputFile, 'w')

#		wb = openpyxl.load_workbook('csv\\goodinfo\\saleMonth\\' + inputFile)
		print(loadFileDir + inputFile)
		wb = openpyxl.load_workbook(loadFileDir + inputFile)
		sheet = wb.worksheets[0]

		isSTOP = False
		irow = 4
		icol = 1

		while not isSTOP :
			theList = []
			for icol in range(1, 18) :
				theValue = sheet.cell(row = irow, column = icol).value
				print(theValue)
			
				if icol == 1 :
					if sheet.cell(row = irow, column = icol).value is None :
#						因為非完整年度會以季表示，可能會出現合併列；目前觀察只會出現在當年度，故需新增判斷排除此情形					
						if irow > 6:
							isSTOP = True
						break
					else :
						theValue = "'" + str(theValue) + "'"
				else :
					if type(theValue) == str :
						if theValue == "-" :
							theValue = NULL_VALUE
				theList.append(theValue)

			print(theList)
	
			if len(theList) > 0 :		
				outfile.write("insert into stocks_per_and_pbr (stock_no, year, \
share_capital, fin_report_score, ann_high_price, \
ann_low_price, ann_end_price, \
ann_avg_price, ann_price_raf, \
ann_price_raf_pct, pes_statistics_eps, \
pes_statistics_high, pes_statistics_low, \
pes_statistics_avg, pbr_statistics_eps, \
pbr_statistics_high, pbr_statistics_low, \
pbr_statistics_avg) values ('" + stockCode + "',")
				outfile.write(",".join([str(_) for _ in theList]))
				outfile.write(");\n")
			insertCnt += 1

			# 預防錯誤，理論上應該不會超過100列
			if irow > 100 :
				print("i=" + str(irow))
				isSTOP = True
			irow += 1
		outfile.close()
	
except IOError as err :
	print('File error : ' + str(err))

print("資料處理完成!! 共 " + str(insertCnt) + " 筆。")
print("[FormatFile.py] 結束執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
