"""
GoodinfoData2DB.CheckFileContent 的 Docstring
CMD : python3 CheckFileContent.py theDirectoryPath theCompareValue thePosition
Example :
-- 1 2 : M1 現金+股票殖利率
(mac) python3 CheckFileContent_v2.py M1 現金+股票殖利率 /Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2/ 
(win-sgo3) python CheckFileContent_v2.py M1 現金+股票殖利率 D:\workspaces\GithubProjects\GoodinfoData2DB\Data\EXCEL\Transfer\dividend\20260215_1_2\
-- 1 4 : M1 除權/息價殖利率
(mac) python3 CheckFileContent_v2.py M1 除權/息價殖利率 /Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Origin/dividend/20260216_1_4/ 
(win-sgo3) python CheckFileContent_v2.py M1 除權/息價殖利率 D:\workspaces\GithubProjects\GoodinfoData2DB\Data\EXCEL\Transfer\dividend\20260216_1_4\

"""
from openpyxl import load_workbook
import os
import sys

#path = "/Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2/"

if len(sys.argv) < 4:
    print("參數不足：theDirectoryPath theCompareValue thePosition")
    sys.exit(1)

thePosition = sys.argv[1]
theDirectoryPath = sys.argv[3]
theCompareValue = sys.argv[2]

if not os.path.isdir(theDirectoryPath):
    print(f"目錄不存在：{theDirectoryPath}")
    sys.exit(1)

files = os.listdir(theDirectoryPath)
#print("files= ", files)
file_count = sum(1 for f in files if os.path.isfile(os.path.join(theDirectoryPath, f)))
print("檔案數量:", file_count)


# 建立一個紀錄檔案
output_file = "___compare_result.txt"

with open(output_file, "w", encoding="utf-8") as out:
    for i in range(file_count):
        file_name = files[i]
        print("file_name: ", file_name)

        wb = load_workbook(theDirectoryPath + file_name)
        ws = wb.active

        value = ws[thePosition].value
        print(value)

        if value == theCompareValue:
            print("OK")
        else:
            print("FAILED")
            # 寫入不符合的檔案名稱與值
            out.write(f"{file_name}: {value}\n")

print(f"比對完成，不符合的結果已寫入 {output_file}")
