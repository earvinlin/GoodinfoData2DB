"""
GoodinfoData2DB.CheckFileContent 的 Docstring
CMD : python3 CheckFileContent.py theDirectoryPath theCompareValue
python3 CheckFileContent.py /Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2/ 現金+股票殖利率

"""

from openpyxl import load_workbook
import os
import sys

#path = "/Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2"
path = "/Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2/"


if len(sys.argv) < 3:
    print("參數不足：theDirectoryPath theCompareValue")
    sys.exit(1)

theDirectoryPath = sys.argv[1]
theCompareValue = sys.argv[2]

if not os.path.isdir(theDirectoryPath):
    print(f"目錄不存在：{theDirectoryPath}")
    sys.exit(1)

# 列出目錄下所有項目
files = os.listdir(theDirectoryPath)

# 計算檔案數量（排除子目錄）
file_count = sum(1 for f in files if os.path.isfile(os.path.join(path, f)))
print("檔案數量:", file_count)


for i in range(0, file_count):    
#    file_name = "0050-dividend-20260215_1_2.xlsx"
    file_name = files[i]
    print("file_name: ", file_name)

    wb = load_workbook(path + file_name)
    ws = wb.active

    value = ws["M1"].value
    print(value)

    if value == theCompareValue :
        print("OK")




"""
import pandas as pd

# ~/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2
path = "~/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/dividend/20260215_1_2/"
file_name = "0050-dividend-20260215_1_2.xlsx"
sheet_name = "0050-dividend-20260215_1_2"
df = pd.read_excel(path + file_name, sheet_name=sheet_name)

row_index = 5
col_index = 2
value = df.iloc[row_index, col_index]
for i in range(0, 10):
    for j in range(0, 10):
        value = df.iloc[i, j]
        print("content[", i, ",", j, "]= ", value)

"""