import sys
import os 
import time
import openpyxl
#import xlrd

NULL_VALUE = "null"
insertCnt = 0 
isFirstLine = True
try:
	print("[FormatFile.py] 開始執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

	if len(sys.argv) < 3 :
		print("You need input one parameter(fmt : 股票代號 資料抓取日期)")
		print("syntax : C:\python FormatStocksBzPerformanceData.py 1101 20220508")
		sys.exit()
	loadFileDir = "Data\\EXCEL\\Transfer\\BzPerformance\\"
	saveFileDir = "Data\\TXT\\BzPerformance\\"
	stockCode = sys.argv[1]
	inputFile = stockCode + "-BzPerformance-" + sys.argv[2] + ".xlsx"
	outputFile = stockCode + ".txt"
	
	outfile = open(saveFileDir + outputFile, 'w')

	wb = openpyxl.load_workbook(loadFileDir + inputFile)
	sheet = wb.worksheets[0]

	isSTOP = False
	irow = 4
	icol = 1

	while not isSTOP :
		theList = []
		for icol in range(1, 22) :
			theValue = sheet.cell(row = irow, column = icol).value
			print(theValue)
			
			if icol == 1 :
				if sheet.cell(row = irow, column = icol).value is None :
#					因為非完整年度會以季表示，可能會出現合併列；目前觀察只會出現在當年度，故需新增判斷排除此情形					
					if irow > 6:
						isSTOP = True
					break
				else :
					theValue = "'" + str(theValue) + "'"
			else :
				if type(theValue) == str :
					if theValue == "-" :
						theValue = NULL_VALUE
			theList.append(theValue)

		print(theList)
		if len(theList) > 0 :
			outfile.write("insert into stocks_bz_performance (stock_no, year, \
share_capital, fin_report_score, ann_stock_end_price, ann_stock_avg_price, \
ann_stock_ud_price, ann_stock_ud_price_pc, profit_revenue, \
profit_gross_profit, profit_business_interest, profit_nonop_g_and_l, \
profit_income_after_taxes, profit_ratio_gp, \
profit_ratio_business_interest, profit_ratio_nonop_g_and_l, \
profit_ratio_income_after_taxes, roe, roa, eps_after_taxes, \
ann_increase_eps, bps) values ('" + stockCode + "',")
			outfile.write(",".join([str(_) for _ in theList]))
			outfile.write(");\n")
		insertCnt += 1

		# 預防錯誤，理論上應該不會超過100列
		if irow > 100 :
			print("i=" + str(irow))
			isSTOP = True
		irow += 1

	outfile.close()
	
except IOError as err :
	print('File error : ' + str(err))

print("資料處理完成!! 共 " + str(insertCnt) + " 筆。")
print("[FormatFile.py] 結束執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))
