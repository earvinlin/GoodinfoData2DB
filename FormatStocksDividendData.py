"""
20220528-2232 格式化產生寫入Stocks_Dividend的SQL Command
"""
import sys
import os 
import time
import openpyxl
import time
import platform
import re
from datetime import datetime
from sqlalchemy import null

NULL_VALUE = "null"
insertCnt = 0 
isFirstLine = True
try:
	print("[FormatStocksDividendData.py] 開始執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

	if len(sys.argv) < 2 :
		print("You need input one parameter(資料日期, fmt : yyyymmdd)")
		print("syntax(windows)    : C:\python FormatStocksDividendData.py 20220517")
		print("syntax(imac/linux) : $python3 FormatStocksDividendData.py 20220517")
		sys.exit()

	theDate = sys.argv[1]

	# 設定檔案存取路徑
	loadFileDir = ""
	saveFileDir = ""
	if platform.system() == "Windows" :
		loadFileDir = "Data\\EXCEL\\Transfer\\dividend\\" + str(theDate) + "\\"
		saveFileDir = "Data\\TXT\\dividend\\" + str(theDate) + "\\"
	else :
		loadFileDir = "./Data/EXCEL/Transfer/dividend/" + str(theDate) + "/"
		saveFileDir = "./Data/TXT/dividend/" + str(theDate) + "/"

	print("loadFile dir: " + loadFileDir)
	print("saveFile dir: " + saveFileDir)

	# 20220522 取得要處理的檔案資料
	files = os.listdir(loadFileDir)
	# 以迴圈處理
	for inputFile in files:
		relativePath = os.path.join(loadFileDir, inputFile)
		if os.path.isfile(relativePath) :
			print("檔案：", inputFile)

		stockCode = inputFile[:4]
		outputFile = stockCode + ".txt"
		print("outfile: " + outputFile)
		outfile = open(saveFileDir + outputFile, 'w')

#		wb = openpyxl.load_workbook('csv\\goodinfo\\saleMonth\\' + inputFile)
		print(loadFileDir + inputFile)
		wb = openpyxl.load_workbook(loadFileDir + inputFile)
		sheet = wb.worksheets[0]

		isSTOP = False
		irow = 6
		icol = 1
		thePrevValue01 = '-'
		thePrevValue13 = '-'

		while not isSTOP :
			theList = []
			theList.append("insert into stocks_dividend \
(stock_no,dividend_year,cash_div_surplus,cash_div_reserve,total_cash_div,\
stock_div_surplus,stock_div_reserve,total_stock_div,total_dividend,\
total_div_cash,total_div_stocks,days_fill_cash,days_fill_stocks,\
stock_price_year,year_high_price,year_low_price,year_avg_price,\
avg_ann_cash_yield,avg_ann_stock_yield,avg_ann_yield,period_of_dividend,\
eps,div_earnings_dis_ratio,alo_earnings_dis_ratio,earnings_dis_ratio) values ('" + stockCode + "'")
			for icol in range(1, 25) :
				theValue = sheet.cell(row = irow, column = icol).value
#				print("icol= " + str(icol) + ", value= " + str(theValue))

				if icol == 1 :
#					股利發放年度：因為如果非年配息(如：季配)會有多筆資料，為了後續查詢作業，本欄資料要另外處理
#					print("thePrevValue01= " + str(thePrevValue01))
					if thePrevValue01 == '-'  :
						thePrevValue01 = str(theValue)
					if str(theValue) == '∟' :
						theValue = thePrevValue01
					if str(theValue) != thePrevValue01 :
						thePrevValue01 = str(theValue)

					if theValue == "累計" :
						isSTOP = True
						theList.pop()
						break
					theList.append(theValue)
				elif icol == 13 :
#					股價年度：因為如果非年配息(如：季配)會有多筆資料，為了後續查詢作業，本欄資料要另外處理
					if thePrevValue13 == '-' :
						thePrevValue13 = str(theValue)
					if str(theValue) == '∟' :
						theValue = thePrevValue13
					if str(theValue) != thePrevValue13 :
						thePrevValue13 = theValue
					if theValue == '-' :
						theValue = NULL_VALUE
					
					theList.append(theValue)
				elif icol == 20 :
# 					股利所屬期間
					theValue = "'" + str(theValue) +"'"
					theList.append(theValue)
				elif type(theValue) == str :
					if theValue == "-" :
						theValue = NULL_VALUE
					else :
						theValue = "'" + theValue +"'"
					theList.append(theValue)
				elif type(theValue) == int :
					theList.append(theValue)
				elif type(theValue) == float :
					theList.append(theValue)
				else :
					theList.append(theValue)
			print(theList)
			if len(theList) > 0 :
				outfile.write(",".join([str(_) for _ in theList]))
				outfile.write(");\n")
			insertCnt += 1

			# 預防錯誤，理論上應該不會超過5000列(1年12筆，100年1200筆；所以約400年的企業才可能有問題)
			if irow > 100 :
				print("i=" + str(irow))
				isSTOP = True
			irow += 1

		outfile.close()
		print("資料處理完成!! 共 " + str(insertCnt) + " 筆。")
		print("[FormatStocksDividendData.py] 結束執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

except IOError as err :
	print('File error : ' + str(err))