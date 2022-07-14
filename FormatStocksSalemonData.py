"""
20220521-2049 調整資料來源為指定之路徑並批次產生sql command
"""
from datetime import datetime
import sys
import os 
import time
import openpyxl
import time
import platform
import re

NULL_VALUE = "null"
insertCnt = 0 
isFirstLine = True
try:
	print("[GetGoodinfoSalemonData.py] 開始執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

	if len(sys.argv) < 2 :
		print("You need input one parameter(資料日期, fmt : yyyymmdd)")
		print("syntax(windows)    : C:\python FormatStocksSalemonData.py 20220517")
		print("syntax(imac/linux) : $python3 FormatStocksSalemonData.py 20220517")
		sys.exit()

	theDate = sys.argv[1]

	# 設定檔案存取路徑
	loadFileDir = ""
	saveFileDir = ""
	if platform.system() == "Windows" :
		loadFileDir = "Data\\EXCEL\\Transfer\\salemon\\" + str(theDate) + "\\"
		saveFileDir = "Data\\TXT\\salemon\\" + str(theDate) + "\\"
	else :
#       iMac下的路徑：「/Users/earvin/workspaces/GithubProjects/GoodinfoData2DB」(目前似乎只能用絕對路徑)
		loadFileDir = "/Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/EXCEL/Transfer/salemon/" + str(theDate) + "/"
		saveFileDir = "/Users/earvin/workspaces/GithubProjects/GoodinfoData2DB/Data/TXT/salemon/" + str(theDate) + "/"

	print("loadFile dir: " + loadFileDir)
	print("saveFile dir: " + saveFileDir)

	# 20220522 取得要處理的檔案資料
	files = os.listdir(loadFileDir)
	# 以迴圈處理
	for inputFile in files:
		relativePath = os.path.join(loadFileDir, inputFile)
		if os.path.isfile(relativePath) :
			print("檔案：", inputFile)

		stockCode = inputFile[:4]
		outputFile = stockCode + ".txt"
		print("outfile: " + outputFile)
		outfile = open(saveFileDir + outputFile, 'w')

#		wb = openpyxl.load_workbook('csv\\goodinfo\\saleMonth\\' + inputFile)
		print(loadFileDir + inputFile)
		wb = openpyxl.load_workbook(loadFileDir + inputFile)
		sheet = wb.worksheets[0]

		isSTOP = False
		irow = 5
		icol = 1

#	theList.append("insert into stocks_sale_month (stock_no,date,open_price_mon,end_price_mon,hgh_price_mon,low_price_mon,ups_downs,ups_downs_p,mon_revenue,mon_increase_in_revenue,ann_ins_revenue,cum_revenue,ann_ins_cum_revenue_p,csd_revenue,mon_ins_csd_revenue_p,ann_ins_csd_revenue_p,cum_csd_revenue,ann_ins_cum_csd_revenue_p) values ('" + stockCode + "'")
#	theSQLCmd = "insert into stocks_sale_month (stock_no, date, open_price_mon, end_price_mon, hgh_price_mon, low_price_mon, ups_downs, ups_downs_p, mon_revenue, mon_increase_in_revenue, ann_ins_revenue, cum_revenue, ann_ins_cum_revenue_p, csd_revenue, mon_ins_csd_revenue_p, ann_ins_csd_revenue_p, cum_csd_revenue, ann_ins_cum_csd_revenue_p) values ('%s', '%d', '%f',  '%f',  '%f',  '%f',  '%f',  '%f',  '%f', '%f', '%f', '%f', '%f', '%f', '%f', '%f', '%f', '%f')"
#	outfile.write(theSQLCmd+"\n")
		while not isSTOP :
			theList = []
			theList.append("insert into stocks_sale_month (stock_no,date,open_price_mon,end_price_mon,hgh_price_mon,low_price_mon,ups_downs,ups_downs_p,mon_revenue,mon_increase_in_revenue,ann_ins_revenue,cum_revenue,ann_ins_cum_revenue_p,csd_revenue,mon_ins_csd_revenue_p,ann_ins_csd_revenue_p,cum_csd_revenue,ann_ins_cum_csd_revenue_p) values ('" + stockCode + "'")

			for icol in range(1, 18) :
				theValue = sheet.cell(row = irow, column = icol).value
#			print(theValue)
#			if sheet.cell(row = irow, column = icol).value == None :
				if icol == 1 and sheet.cell(row = irow, column = icol).value == None :
					isSTOP = True
					theList.pop()
					break
				elif type(theValue) == str :
					if theValue == "-" :
						theValue = NULL_VALUE
					theList.append(theValue)
				elif type(theValue) == int :
					theList.append(theValue)
				elif type(theValue) == float :
					theList.append(theValue)
				elif type(theValue) == datetime :
					theList.append(str(theValue.strftime("%Y%m%d")))
				else :
					theList.append(theValue)
#			print(theList)
			if len(theList) > 0 :
				outfile.write(", ".join([str(_) for _ in theList]))
				outfile.write(");\n")
			insertCnt += 1

			# 預防錯誤，理論上應該不會超過5000列(1年12筆，100年1200筆；所以約400年的企業才可能有問題)
			if irow > 5000 :
				print("i=" + str(irow))
				isSTOP = True
			irow += 1

		outfile.close()
		print("資料處理完成!! 共 " + str(insertCnt) + " 筆。")
		print("[GetGoodinfoSalemonData.py] 結束執行時間：" + time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()))

except IOError as err :
	print('File error : ' + str(err))
